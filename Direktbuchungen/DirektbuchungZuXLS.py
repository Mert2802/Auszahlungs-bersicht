import os
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# =========================================================
# EINSTELLUNGEN
# =========================================================
CONFIG_FILE = "Unterkunfts_IDs.xlsx"
FIRMEN_ADRESSE_FALLBACK = "Mülheimer Str. 15" 

def load_id_mapping():
    mapping = {}
    if not os.path.exists(CONFIG_FILE):
        print(f"ACHTUNG: '{CONFIG_FILE}' fehlt!")
        return {}
    try:
        df = pd.read_excel(CONFIG_FILE, dtype=str)
        if df.empty: return {}
        df.columns.values[0] = "adresse"
        df.columns.values[1] = "id"
        for index, row in df.iterrows():
            addr = str(row['adresse']).strip()
            uid = str(row['id']).strip()
            if addr.lower() != "nan" and uid.lower() != "nan":
                if uid.endswith(".0"): uid = uid[:-2]
                mapping[addr] = uid
        print(f"ID-Liste geladen: {len(mapping)} Einträge.")
        return mapping
    except Exception as e:
        print(f"Fehler Excel: {e}")
        return {}

ID_MAPPING = load_id_mapping()

def parse_german_float(s):
    if not s: return 0.0
    match = re.search(r"([\d.]+,\d{2})", s) 
    if match: s = match.group(1)
    else: s = s.replace('€', '').replace('%', '').strip()
    s = s.replace('.', '').replace(',', '.')
    try: return float(s)
    except: return 0.0

def normalize(s):
    if not s: return ""
    s = str(s).lower().replace(" ", "").replace(".", "").replace("-", "")
    s = s.replace("strasse", "str").replace("straße", "str")
    return s

def find_address_in_line(line_text):
    line_norm = normalize(line_text)
    for addr_key in sorted(ID_MAPPING.keys(), key=len, reverse=True):
        if normalize(addr_key) in line_norm:
            return addr_key
    return None

folder_path = os.getcwd()
pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]
extracted_data = []

print(f"{len(pdf_files)} PDF-Dateien gefunden. Starte V23 (Steuer = 5% von BmGl)...")

for filename in pdf_files:
    file_path = os.path.join(folder_path, filename)
    print(f"Analysiere: {filename}...")
    
    buchungen_in_datei = [] 
    total_cleaning_netto_file = 0.0
    cleaning_vat_rate = 0.0
    steuer_aktiv = False
    
    single_endbetrag = 0.0
    single_reinigung_brutto = 0.0
    single_steuer = 0.0
    
    try:
        with pdfplumber.open(file_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text: full_text += page_text + "\n"

            # Zeilen entwirren
            full_text = full_text.replace("PrimeTimeSuite", "\nPrimeTimeSuite")
            full_text = full_text.replace("Prime TimeSuite", "\nPrime TimeSuite")
            lines = full_text.split('\n')

            # --- SCAN PHASE ---
            for line in lines:
                
                # Endbetrag
                match_end = re.search(r"Endbetrag:.*?([\d.,]+)", line)
                if match_end: single_endbetrag = parse_german_float(match_end.group(1))

                # Buchungen
                found_addr = find_address_in_line(line)
                betraege = re.findall(r"[\d.]+,\d{2}", line)
                
                if found_addr and betraege:
                    preis = parse_german_float(betraege[-1])
                    if preis > 0 and not re.search(r"Rein\w*", line) and "Beherbergungssteuer" not in line:
                        buchungen_in_datei.append({
                            "Strasse": found_addr,
                            "Netto_Accom": preis
                        })

                # Reinigung
                if re.search(r"Rein\w*", line, re.IGNORECASE):
                    betraege = re.findall(r"[\d.]+,\d{2}", line)
                    vals = [parse_german_float(b) for b in betraege]
                    vals = [v for v in vals if v > 0]
                    if vals:
                        line_cleaning_total = max(vals)
                        total_cleaning_netto_file += line_cleaning_total
                        match_prz = re.search(r"(\d{1,2}(?:,\d{1,2})?)\s*%", line)
                        if match_prz: cleaning_vat_rate = parse_german_float(match_prz.group(1))
                        if cleaning_vat_rate == 0: cleaning_vat_rate = 19.0
                        single_reinigung_brutto = line_cleaning_total * (1 + cleaning_vat_rate/100)

                # Steuer
                if "Beherbergungssteuer" in line:
                    steuer_aktiv = True
                    if not "5%" in line:
                        betraege = re.findall(r"[\d.]+,\d{2}", line)
                        if betraege: single_steuer = parse_german_float(betraege[-1])

            # --- LOGIK PHASE ---
            num_bookings = len(buchungen_in_datei)
            
            # FALL A: Einzel-Rechnung
            if num_bookings <= 1:
                final_addr = buchungen_in_datei[0]["Strasse"] if num_bookings == 1 else "Unbekannt"
                
                if single_endbetrag == 0 and num_bookings == 1:
                     single_endbetrag = (buchungen_in_datei[0]["Netto_Accom"] * 1.07) + single_reinigung_brutto + single_steuer
                
                # Wir ziehen die Steuer vom Endbetrag ab, um den "Brutto-Umsatz ohne CityTax" zu erhalten.
                # Damit vermeiden wir, dass wir später Steuer auf Steuer zahlen.
                brutto_einkunft_final = single_endbetrag - single_steuer
                
                extracted_data.append({
                    "Strasse": final_addr,
                    "Bruttoeinkünfte": brutto_einkunft_final,
                    "Reinigungskosten": single_reinigung_brutto,
                    "BmGl Beherbergungssteuer": 0.0, # Wird unten berechnet
                    "Beherbergungssteuer 5% von SP3": 0.0 # Wird unten berechnet
                })

            # FALL B: Monster-Rechnung
            else:
                if cleaning_vat_rate == 0: cleaning_vat_rate = 19.0
                total_cleaning_brutto = total_cleaning_netto_file * (1 + cleaning_vat_rate / 100)
                cleaning_per_booking = total_cleaning_brutto / num_bookings
                
                for b in buchungen_in_datei:
                    rent_gross = b["Netto_Accom"] * 1.07
                    # Wir ignorieren hier die PDF-Steuer erst mal und berechnen sie unten neu
                    revenue_gross = rent_gross + cleaning_per_booking
                    
                    extracted_data.append({
                        "Strasse": b["Strasse"],
                        "Bruttoeinkünfte": revenue_gross,
                        "Reinigungskosten": cleaning_per_booking,
                        "BmGl Beherbergungssteuer": 0.0, # Wird unten berechnet
                        "Beherbergungssteuer 5% von SP3": 0.0 # Wird unten berechnet
                    })

    except Exception as e:
        print(f"Fehler: {e}")

# =========================================================
# NACHBEARBEITUNG (BmGl & Steuer)
# =========================================================
for entry in extracted_data:
    # 1. BmGl berechnen (Brutto - Reinigung)
    entry["BmGl Beherbergungssteuer"] = entry["Bruttoeinkünfte"] - entry["Reinigungskosten"]
    
    # 2. Steuer berechnen (5% von BmGl)
    # Hier wird jetzt strikt nach Tabelle gerechnet!
    if entry["BmGl Beherbergungssteuer"] > 0:
        entry["Beherbergungssteuer 5% von SP3"] = entry["BmGl Beherbergungssteuer"] * 0.05
    else:
        entry["Beherbergungssteuer 5% von SP3"] = 0.0

# --- EXCEL ERSTELLUNG ---
if extracted_data:
    df = pd.DataFrame(extracted_data)
    df_sum = df.groupby("Strasse").sum().reset_index()
    
    def find_id(addr):
        norm_addr = normalize(addr)
        for key, val in ID_MAPPING.items():
            if normalize(key) in norm_addr: return val
        return "ID FEHLT"

    df_sum.insert(0, "BeherbergungsID", df_sum["Strasse"].apply(find_id))
    
    numeric_cols = ["Bruttoeinkünfte", "Reinigungskosten", "BmGl Beherbergungssteuer", "Beherbergungssteuer 5% von SP3"]
    totals = df_sum[numeric_cols].sum()
    total_row = pd.DataFrame([["GESAMT", ""] + totals.tolist()], columns=["BeherbergungsID", "Strasse"] + numeric_cols)
    df_sum = pd.concat([df_sum, total_row], ignore_index=True)

    output_filename = "Airbnb_Style_Report.xlsx"
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df_sum.to_excel(writer, index=False, sheet_name='Buchungen')
        ws = writer.sheets['Buchungen']
        
        purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for cell in ws[1]:
            cell.fill = purple_fill; cell.font = white_font; cell.alignment = Alignment('center', 'center'); cell.border = thin_border
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                if cell.col_idx >= 3: cell.number_format = '#,##0.00 €'
                if cell.row == ws.max_row: cell.font = bold_font

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value)) for c in col) + 4

    print("\n" + "="*40)
    print("FERTIG! Alles ist jetzt konsistent:")
    print("1. BmGl = Brutto - Reinigung")
    print("2. Steuer = BmGl * 0,05")
    print("="*40)
else:
    print("Keine Daten gefunden.")

input("Drücke ENTER zum Beenden...")